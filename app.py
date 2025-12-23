from flask import Flask, request, jsonify
import requests
import base64
import json
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename
import logging
from PIL import Image
import io
import time
import zipfile
import tempfile
import shutil

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configurações
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
TEMPLATE_FILE_PATH = "./Registro Ponto.XLSX"
# Alias para retrocompatibilidade
EXCEL_FILE_PATH = TEMPLATE_FILE_PATH
GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-3-pro-preview:generateContent"

def get_extraction_prompt():
    """Prompt otimizado para extração de horários"""
    return """Analise a imagem de cartão-ponto anexada e devolva **somente** um objeto JSON.

OBJETIVO
• Cada chave deve ser o dia do mês com dois dígitos (ex.: "01", "15").
• Cada valor deve ser um array contendo **todos** os horários encontrados nesse dia no formato HH:MM.
• Exemplo de retorno desejado:
  {
    "16": ["09:04", "13:00", "14:00", "18:02"],
    "17": ["08:56", "13:00", "14:00", "18:01"],
    "29": ["10:03", "12:00", "15:00*", "14:54"]
  }

REGRAS DE EXTRAÇÃO
1. Ignore linhas ou dias sem horário legível.
2. Ignore qualquer texto que não seja horário no formato HH:MM.
3. Se um horário estiver rasurado, tachado, parcialmente ilegível ou escrito fora do padrão, inclua-o mesmo assim **adicionando um asterisco "*" ao final** (ex.: "09:30*").
4. Não inclua campos extras, títulos ou comentários — apenas o objeto JSON resultante.

CHAIN OF THOUGHT (INTERNO — NÃO EXPOR)
1. Detecte a grade/table do cartão.
2. Isole cada célula de horário.
3. Reconheça o texto (OCR).
4. **Para cada horário potencialmente rasurado:** diminua a velocidade, aplique múltiplos limiares de contraste/brilho, compare hipóteses e vote internamente antes de decidir pelo valor final; se a incerteza permanecer, marque com "*".
5. Valide se corresponde a HH:MM; em caso de rasura ou escrita manual, acrescente "*".
8. Verifique se a linha do dia está destacada por marca-texto (verde). Se sim, inclua no JSON a chave "painted": true; caso contrário "painted": false.
9. A saída final deve ter, para cada dia, objeto com "times" (array) e "painted" (boolean). Exemplo:
{
  "20": { "times": ["1018", "1200", "1302", "1649"], "painted": true },
  "21": { "times": ["1005*", "1304", "1403", "1857"], "painted": true },
  "22": { "times": ["1027", "1304", "1405", "1900"], "painted": false }
}
6. Agrupe por dia preservando a ordem em que os horários aparecem.
7. Gere o JSON e responda apenas com ele, sem revelar este raciocínio."""

def process_with_gemini(image_b64):
    """Processa imagem com Google Gemini"""
    try:
        headers = {
            "Content-Type": "application/json",
            "X-goog-api-key": GEMINI_API_KEY
        }
        
        payload = {
            "contents": [
                {
                    "role": "user",
                    "parts": [
                        {"text": get_extraction_prompt()},
                        {
                            "inline_data": {
                                "mime_type": "image/jpeg",
                                "data": image_b64
                            }
                        }
                    ]
                }
            ],
            "generationConfig": {
                "response_mime_type": "application/json"
            }
        }
        
        logger.info("Enviando requisição para Gemini...")
        retries = 3
        delay = 5  # segundos
        for attempt in range(1, retries + 1):
                try:
                    response = requests.post(GEMINI_URL, headers=headers, json=payload, timeout=60)
                    response.raise_for_status()
                    result = response.json()
                    if "candidates" in result and len(result["candidates"]) > 0:
                        content = result["candidates"][0]["content"]["parts"][0]["text"]
                        logger.info("Resposta recebida do Gemini")
                        return content
                    else:
                        raise Exception("Resposta inválida do Gemini")
                except requests.exceptions.RequestException as e:
                    logger.warning(f"Tentativa {attempt}/{retries} falhou: {e}")
                    if attempt == retries:
                        raise Exception(f"Falha no processamento da imagem após {retries} tentativas: {e}")
                    logger.info(f"Aguardando {delay}s para retry...")
                    time.sleep(delay)
                    delay *= 2
            
    except Exception as e:
        logger.error(f"Erro na integração com Gemini: {str(e)}")
        raise Exception(f"Falha no processamento da imagem: {str(e)}")

def update_excel(time_entries):
    """Atualiza planilha Excel com horários"""
    try:
        logger.info(f"Atualizando planilha: {EXCEL_FILE_PATH}")
        
        # Verifica se arquivo existe
        if not os.path.exists(EXCEL_FILE_PATH):
            logger.warning("Arquivo Excel não encontrado, criando novo...")
            create_new_workbook()
        
        # Carrega workbook
        workbook = load_workbook(EXCEL_FILE_PATH)
        
        # Seleciona planilha
        if workbook.worksheets:
            worksheet = workbook.active
        else:
            worksheet = workbook.create_sheet("Registro Ponto")
        
        # Encontra próxima linha disponível
        next_row = worksheet.max_row + 1 if worksheet.max_row else 1
        
        # Adiciona cabeçalhos se necessário
        if next_row == 1:
            headers = ["Data", "Dia", "Horário", "Status", "Timestamp"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=next_row, column=col, value=header)
            next_row = 2
        
        # Insere dados
        rows_added = 0
        for day, times in time_entries.items():
            for time_entry in times:
                current_date = f"{datetime.now().year}-{datetime.now().month:02d}-{day}"
                
                worksheet.cell(row=next_row, column=1, value=current_date)    # Data
                worksheet.cell(row=next_row, column=2, value=int(day))        # Dia
                worksheet.cell(row=next_row, column=3, value=time_entry)      # Horário
                worksheet.cell(row=next_row, column=4, value="Processado")    # Status
                worksheet.cell(row=next_row, column=5, value=datetime.now())  # Timestamp
                
                next_row += 1
                rows_added += 1
        
        # Salva arquivo
        workbook.save(EXCEL_FILE_PATH)
        logger.info(f"Planilha atualizada com {rows_added} registros")
        
        return {"updated": True, "rows_added": rows_added}
        
    except Exception as e:
        logger.error(f"Erro ao atualizar planilha: {str(e)}")
        return {"updated": False, "error": str(e)}

def create_new_workbook():
    """Cria novo workbook Excel"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Registro Ponto"
    
    # Cabeçalhos
    headers = ["Data", "Dia", "Horário", "Status", "Timestamp"]
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    workbook.save(EXCEL_FILE_PATH)
    logger.info(f"Novo workbook criado: {EXCEL_FILE_PATH}")

@app.route('/')
def home():
    """Página inicial"""
    return jsonify({
        "message": "🕘 API de Cartão-Ponto com Flask",
        "version": "2.0.0",
        "endpoints": {
            "POST /process": "Processar cartão-ponto individual",
            "POST /process-month": "Processar mês completo (2 quinzenas) - multipart/form-data",
            "POST /process-month-base64": "Processar mês completo (2 quinzenas) - JSON com base64",
            "GET /health": "Verificar saúde da API",
            "GET /status": "Status da planilha"
        },
        "usage": {
            "/process": "curl -F 'file=@cartao.jpg' http://localhost:5001/process",
            "/process-month": "curl -F 'file1=@quinzena1.jpg' -F 'file2=@quinzena2.jpg' http://localhost:5001/process-month",
            "/process-month-base64": "curl -H 'Content-Type: application/json' -d '{\"image1_base64\":\"...\", \"image2_base64\":\"...\"}' http://localhost:5001/process-month-base64"
        }
    })

@app.route('/health')
def health():
    """Verificação de saúde"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "services": {
            "gemini": "ok",
            "excel": "ok" if os.path.exists(EXCEL_FILE_PATH) else "file_not_found"
        }
    })

@app.route('/status')
def status():
    """Status da planilha"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({
                "file_exists": False,
                "file_path": EXCEL_FILE_PATH,
                "error": "Arquivo não encontrado"
            })
        
        stat = os.stat(EXCEL_FILE_PATH)
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active
        
        return jsonify({
            "file_exists": True,
            "file_path": EXCEL_FILE_PATH,
            "last_modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "total_rows": worksheet.max_row if worksheet.max_row > 1 else 0,
            "file_size": stat.st_size
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/process-month', methods=['POST'])
def process_full_month():
    """
    Processa 2 imagens de cartão-ponto (1ª e 2ª quinzena) e atualiza planilha Excel
    
    Args:
        file1: Arquivo da 1ª quinzena (JPEG, PNG)
        file2: Arquivo da 2ª quinzena (JPEG, PNG)
        
    Returns:
        ProcessResponse: Resultado do processamento do mês completo
    """
    try:
        # Verifica se ambos os arquivos foram enviados
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({"error": "Envie os dois arquivos: file1 (1ª quinzena) e file2 (2ª quinzena)"}), 400
        
        file1 = request.files['file1']  # 1ª quinzena
        file2 = request.files['file2']  # 2ª quinzena
        
        if file1.filename == '' or file2.filename == '':
            return jsonify({"error": "Ambos os arquivos devem ser selecionados"}), 400
        
        # Verifica tipos de arquivo
        for file in [file1, file2]:
            content_type = file.content_type or 'application/octet-stream'
            if not content_type.startswith('image/') and not file.filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                return jsonify({"error": f"Arquivo {file.filename} deve ser uma imagem (JPG, PNG)"}), 400
        
        logger.info(f"Processando mês completo: {file1.filename} + {file2.filename}")
        
        # Lê conteúdo das imagens
        image1_content = file1.read()
        image2_content = file2.read()
        
        # Combina as imagens lado a lado
        combined_image_b64 = combine_images_side_by_side(image1_content, image2_content)
        
        # Processa com Gemini
        gemini_result = process_with_gemini(combined_image_b64)
        
        # Parse do JSON
        try:
            time_entries = json.loads(gemini_result)
        except json.JSONDecodeError as e:
            return jsonify({"error": f"Resposta inválida do Gemini: {str(e)}"}), 500
        
        # Atualiza planilha Excel com estrutura correta
        excel_result = update_excel_structured(time_entries, image1_content, image2_content)
        
        logger.info(f"Processamento concluído: {len(time_entries)} dias processados")
        
        return jsonify({
            "success": True,
            "message": "Mês completo processado com sucesso! ✅",
            "time_entries": time_entries,
            "excel_updated": excel_result["updated"],
            "file_base64": excel_result.get("file_base64"),
            "summary": {
                "total_days": len(time_entries),
                "total_entries": sum(len(entries) for entries in time_entries.values()),
                "file1_name": file1.filename,
                "file2_name": file2.filename,
                "rows_updated": excel_result.get("rows_updated", 0),
                "output_file": excel_result.get("output_file"),
                "image_inserted": excel_result.get("image_inserted", False),
                "capa_updated": excel_result.get("capa_updated", False)
            },
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Erro interno: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }), 500

@app.route('/process-month-base64', methods=['POST'])
def process_full_month_base64():
    """
    Processa 2 imagens de cartão-ponto em base64 (1ª e 2ª quinzena) e retorna planilha Excel
    
    Args:
        JSON payload com:
        - image1_base64: String base64 da 1ª quinzena
        - image2_base64: String base64 da 2ª quinzena
        
    Returns:
        ProcessResponse: Resultado do processamento do mês completo com Excel em base64
    """
    try:
        # Verifica se é JSON
        if not request.is_json:
            return jsonify({"error": "Content-Type deve ser application/json"}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({"error": "Payload JSON inválido"}), 400
        
        # Verifica se ambas as imagens base64 foram enviadas
        if 'image1_base64' not in data or 'image2_base64' not in data:
            return jsonify({"error": "Envie as duas imagens: image1_base64 (1ª quinzena) e image2_base64 (2ª quinzena)"}), 400
        
        image1_b64 = data['image1_base64']
        image2_b64 = data['image2_base64']
        
        if not image1_b64 or not image2_b64:
            return jsonify({"error": "Ambas as imagens base64 devem ser fornecidas"}), 400
        
        logger.info("Processando mês completo via base64")
        
        # Decodifica as imagens base64 para bytes
        try:
            # Remove prefixo data:image se existir
            if image1_b64.startswith('data:image'):
                image1_b64 = image1_b64.split(',')[1]
            if image2_b64.startswith('data:image'):
                image2_b64 = image2_b64.split(',')[1]
                
            image1_content = base64.b64decode(image1_b64)
            image2_content = base64.b64decode(image2_b64)
        except Exception as e:
            return jsonify({"error": f"Erro ao decodificar base64: {str(e)}"}), 400
        
        # Combina as imagens lado a lado
        combined_image_b64 = combine_images_side_by_side(image1_content, image2_content)
        
        # Processa com Gemini
        gemini_result = process_with_gemini(combined_image_b64)
        
        # Parse do JSON
        try:
            time_entries = json.loads(gemini_result)
        except json.JSONDecodeError as e:
            return jsonify({"error": f"Resposta inválida do Gemini: {str(e)}"}), 500
        
        # Atualiza planilha Excel com estrutura correta
        excel_result = update_excel_structured(time_entries, image1_content, image2_content)
        
        logger.info(f"Processamento base64 concluído: {len(time_entries)} dias processados")
        
        return jsonify({
            "success": True,
            "message": "Mês completo processado com sucesso via base64! ✅",
            "time_entries": time_entries,
            "excel_updated": excel_result["updated"],
            "file_base64": excel_result.get("file_base64"),
            "summary": {
                "total_days": len(time_entries),
                "total_entries": sum(len(entries) for entries in time_entries.values()),
                "input_method": "base64",
                "rows_updated": excel_result.get("rows_updated", 0),
                "output_file": excel_result.get("output_file"),
                "image_inserted": excel_result.get("image_inserted", False),
                "capa_updated": excel_result.get("capa_updated", False)
            },
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro no processamento base64: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Erro interno: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }), 500

@app.route('/process', methods=['POST'])
def process_timecard():
    """Processa cartão-ponto"""
    try:
        # Verifica se arquivo foi enviado
        if 'file' not in request.files:
            return jsonify({"error": "Nenhum arquivo enviado"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Nenhum arquivo selecionado"}), 400
        
        # Verifica tipo de arquivo
        content_type = file.content_type or 'application/octet-stream'
        if not content_type.startswith('image/') and not file.filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            return jsonify({"error": "Arquivo deve ser uma imagem (JPG, PNG)"}), 400
        
        logger.info(f"Processando arquivo: {file.filename}")
        
        # Lê conteúdo da imagem
        image_content = file.read()
        image_b64 = base64.b64encode(image_content).decode('utf-8')
        
        # Processa com Gemini
        gemini_result = process_with_gemini(image_b64)
        
        # Parse do JSON
        try:
            time_entries = json.loads(gemini_result)
        except json.JSONDecodeError as e:
            return jsonify({"error": f"Resposta inválida do Gemini: {str(e)}"}), 500
        
        # Atualiza Excel
        excel_result = update_excel(time_entries)
        
        logger.info(f"Processamento concluído: {len(time_entries)} dias processados")
        
        return jsonify({
            "success": True,
            "message": "Cartão-ponto processado com sucesso! ✅",
            "time_entries": time_entries,
            "excel_updated": excel_result["updated"],
            "summary": {
                "total_days": len(time_entries),
                "total_entries": sum(len(entries) for entries in time_entries.values()),
                "file_name": file.filename,
                "rows_added": excel_result.get("rows_added", 0)
            },
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Erro no processamento: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Erro interno: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }), 500

def combine_images_side_by_side(image1_bytes, image2_bytes):
    """Combina duas imagens lado a lado e retorna em base64"""
    try:
        # Abre as imagens
        img1 = Image.open(io.BytesIO(image1_bytes))
        img2 = Image.open(io.BytesIO(image2_bytes))
        
        # Redimensiona para mesma altura se necessário
        if img1.height != img2.height:
            new_height = min(img1.height, img2.height)
            img1 = img1.resize((int(img1.width * new_height / img1.height), new_height))
            img2 = img2.resize((int(img2.width * new_height / img2.height), new_height))
        
        # Cria imagem combinada
        combined_width = img1.width + img2.width
        combined_height = img1.height
        
        combined_img = Image.new('RGB', (combined_width, combined_height))
        combined_img.paste(img1, (0, 0))
        combined_img.paste(img2, (img1.width, 0))
        
        # Converte para base64
        buffered = io.BytesIO()
        combined_img.save(buffered, format="JPEG", quality=85)
        img_b64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
        
        logger.info(f"Imagens combinadas: {img1.size} + {img2.size} → {combined_img.size}")
        return img_b64
        
    except Exception as e:
        logger.error(f"Erro ao combinar imagens: {str(e)}")
        raise

def update_excel_structured(time_entries, image1_bytes, image2_bytes):
    """Atualiza planilha Excel com estrutura correta"""
    try:
        # Gera nome do arquivo de saída (apenas para referência)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Registro_Ponto_{timestamp}.xlsx"

        # Carrega template diretamente em memória
        logger.info(f"📋 Carregando template em memória: {TEMPLATE_FILE_PATH}")
        workbook = load_workbook(TEMPLATE_FILE_PATH)
        
        # 1. Atualiza data na CAPA (C4) - primeiro dia do mês anterior
        if 'CAPA' in workbook.sheetnames:
            capa_sheet = workbook['CAPA']
            today = datetime.now()
            first_day_prev_month = today.replace(day=1) - timedelta(days=1)
            first_day_prev_month = first_day_prev_month.replace(day=1)
            capa_sheet['C4'] = first_day_prev_month
            logger.info(f"CAPA C4 atualizada: {first_day_prev_month.strftime('%d/%m/%Y')}")
        
        # 2. Usa aba FUNCIONÁRIO
        if 'FUNCIONÁRIO' not in workbook.sheetnames:
            return {"updated": False, "error": "Aba FUNCIONÁRIO não encontrada"}
        
        worksheet = workbook['FUNCIONÁRIO']
        
        # 3. Insere horários nas células D7:G37
        rows_updated = 0
        for day_str, entry in time_entries.items():
            # entry pode ser list ou dict
            if isinstance(entry, dict):
                times = entry.get('times', [])
                painted_flag = entry.get('painted', False)
            else:
                times = entry
                painted_flag = False
            try:
                day_num = int(day_str)
                if 1 <= day_num <= 31:
                    row = 6 + day_num  # Dia 1 → linha 7, Dia 2 → linha 8, etc.
                    
                    # Limpa a linha primeiro
                    for col in ['D', 'E', 'F', 'G']:
                        worksheet[f"{col}{row}"] = ""
                    
                    # Insere os horários (máximo 4)
                    for i, time_str in enumerate(times[:4]):
                        col = chr(68 + i)  # D, E, F, G
                        # Remove dois-pontos mantendo asterisco
                        cleaned = time_str.replace(':', '')
                        worksheet[f"{col}{row}"] = cleaned
                    # Aplica fill se pintado
                    if painted_flag:
                        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        for col_letter in ['D','E','F','G']:
                            worksheet[f"{col_letter}{row}"].fill = fill
                    
                    rows_updated += 1
                    logger.info(f"Dia {day_num} → Linha {row}: {times}")
                    
            except ValueError:
                logger.warning(f"Dia inválido ignorado: {day_str}")
        
        # 4. Insere imagem combinada na célula U15
        image_inserted = False
        try:
            # Combina imagens novamente para inserir na planilha
            combined_img_bytes = combine_images_for_excel(image1_bytes, image2_bytes)
            
            # Cria objeto PIL Image a partir dos bytes
            pil_img = Image.open(io.BytesIO(combined_img_bytes))
            
            # Converte para objeto de imagem do openpyxl
            excel_img = ExcelImage(pil_img)
            
            # Redimensiona a imagem para caber bem
            # Mantém aspect ratio e aumenta 200%
            factor = 2  # 200%
            new_width = int(pil_img.width * factor)
            new_height = int(pil_img.height * factor)
            excel_img.width = new_width
            excel_img.height = new_height
            
            worksheet.add_image(excel_img, 'W6')
            image_inserted = True
            logger.info("Imagem inserida na célula U15")
            
        except Exception as e:
            logger.error(f"Erro ao inserir imagem: {str(e)}")
        
        # 5. Converte workbook para base64 e preserva calcChain.xml
        file_base64 = None
        try:
            # Salva workbook em buffer de memória
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            file_bytes = buffer.getvalue()

            # Preserva calcChain.xml do template original
            file_bytes = preserve_calc_chain(TEMPLATE_FILE_PATH, file_bytes)

            file_base64 = base64.b64encode(file_bytes).decode('utf-8')
            logger.info("Arquivo convertido para base64 com calcChain preservado")
        except Exception as e:
            logger.error(f"Erro ao converter para base64: {str(e)}")
        
        logger.info(f"Planilha atualizada: {rows_updated} dias, imagem: {image_inserted}")
        
        return {
            "updated": True,
            "rows_updated": rows_updated,
            "output_file": output_filename,  # Nome do arquivo (sem caminho físico)
            "file_base64": file_base64,
            "image_inserted": image_inserted,
            "capa_updated": True
        }
        
    except Exception as e:
        logger.error(f"Erro ao atualizar planilha: {str(e)}")
        return {"updated": False, "error": str(e)}

def preserve_calc_chain(original_xlsx_path, processed_xlsx_bytes):
    """
    Preserva o calcChain.xml do arquivo original no arquivo processado.

    O openpyxl remove o calcChain.xml ao salvar, o que pode causar problemas
    com fórmulas que dependem de outras células (como detecção de feriados).
    Esta função reinsere o calcChain.xml para garantir que os cálculos
    funcionem corretamente quando o arquivo for aberto no Excel.
    """
    try:
        # 1. Extrai calcChain.xml do original
        with zipfile.ZipFile(original_xlsx_path, 'r') as z_orig:
            if 'xl/calcChain.xml' not in z_orig.namelist():
                logger.warning("Arquivo original não possui calcChain.xml")
                return processed_xlsx_bytes
            calc_chain = z_orig.read('xl/calcChain.xml')

        logger.info(f"calcChain.xml extraído: {len(calc_chain)} bytes")

        # 2. Cria arquivo temporário com o processado
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_processed:
            tmp_processed.write(processed_xlsx_bytes)
            tmp_processed_path = tmp_processed.name

        # 3. Adiciona calcChain.xml ao arquivo processado
        with zipfile.ZipFile(tmp_processed_path, 'a') as z_proc:
            z_proc.writestr('xl/calcChain.xml', calc_chain)

        # 4. Atualiza [Content_Types].xml para incluir referência ao calcChain
        with zipfile.ZipFile(tmp_processed_path, 'r') as z_proc:
            content_types = z_proc.read('[Content_Types].xml').decode('utf-8')

        if 'calcChain' not in content_types:
            logger.info("Atualizando [Content_Types].xml para incluir calcChain")
            calc_ref = '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>'
            content_types = content_types.replace('</Types>', calc_ref + '</Types>')

            # Reescreve o arquivo com o Content_Types atualizado
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_final:
                tmp_final_path = tmp_final.name

            with zipfile.ZipFile(tmp_processed_path, 'r') as z_in:
                with zipfile.ZipFile(tmp_final_path, 'w', zipfile.ZIP_DEFLATED) as z_out:
                    for item in z_in.namelist():
                        if item == '[Content_Types].xml':
                            z_out.writestr(item, content_types.encode('utf-8'))
                        else:
                            z_out.writestr(item, z_in.read(item))

            # Lê o arquivo final
            with open(tmp_final_path, 'rb') as f:
                result_bytes = f.read()

            # Limpa arquivos temporários
            os.unlink(tmp_processed_path)
            os.unlink(tmp_final_path)
        else:
            # Lê o arquivo com calcChain adicionado
            with open(tmp_processed_path, 'rb') as f:
                result_bytes = f.read()
            os.unlink(tmp_processed_path)

        logger.info("calcChain.xml preservado com sucesso")
        return result_bytes

    except Exception as e:
        logger.error(f"Erro ao preservar calcChain: {str(e)}")
        # Em caso de erro, retorna o arquivo original processado
        return processed_xlsx_bytes


def combine_images_for_excel(image1_bytes, image2_bytes):
    """Combina imagens especificamente para inserção no Excel"""
    try:
        img1 = Image.open(io.BytesIO(image1_bytes))
        img2 = Image.open(io.BytesIO(image2_bytes))
        
        # Redimensiona para tamanho otimizado para Excel
        max_width = 600  # largura máxima para cada imagem
        
        if img1.width > max_width:
            img1 = img1.resize((max_width, int(img1.height * max_width / img1.width)))
        if img2.width > max_width:
            img2 = img2.resize((max_width, int(img2.height * max_width / img2.width)))
        
        # Ajusta altura
        if img1.height != img2.height:
            new_height = min(img1.height, img2.height)
            img1 = img1.resize((int(img1.width * new_height / img1.height), new_height))
            img2 = img2.resize((int(img2.width * new_height / img2.height), new_height))
        
        # Combina
        combined_width = img1.width + img2.width
        combined_img = Image.new('RGB', (combined_width, img1.height))
        combined_img.paste(img1, (0, 0))
        combined_img.paste(img2, (img1.width, 0))
        
        # Retorna bytes
        buffered = io.BytesIO()
        combined_img.save(buffered, format="JPEG", quality=90)
        return buffered.getvalue()
        
    except Exception as e:
        logger.error(f"Erro ao combinar para Excel: {str(e)}")
        raise

if __name__ == '__main__':
    print("🚀 Iniciando API de Cartão-Ponto...")
    print("📍 Acesse: http://localhost:5001")
    print("📋 Teste: http://localhost:5001/health")
    print("📄 Novo endpoint: POST /process-month (2 imagens)")
    app.run(debug=True, host='0.0.0.0', port=5001)
